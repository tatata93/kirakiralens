from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sqlite3
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from ..domain import SurfaceSpec
from .database import connect, initialize_schema, insert_surfaces, raw_json


NUMBER_PATTERN = re.compile(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)")

SHAPE_MAP = {
    "Plano-Convex Lens": "plano_convex",
    "Plano-Concave Lens": "plano_concave",
    "Double-Convex Lens": "double_convex",
    "Double-Concave Lens": "double_concave",
    "Achromatic Lens": "achromatic_doublet",
}


@dataclass(slots=True)
class ParsedProduct:
    part_number: str
    title: str
    product_type: str
    shape: str
    outer_diameter_mm: float | None
    clear_aperture_mm: float | None
    effective_focal_length_mm: float | None
    back_focal_length_mm: float | None
    coating: str
    wavelength_min_nm: float | None
    wavelength_max_nm: float | None
    reference_wavelength_nm: float | None
    surfaces: list[SurfaceSpec]
    designable: bool
    raw: dict[str, Any]
    missing_fields: list[str] = field(default_factory=list)


@dataclass(slots=True)
class ImportReport:
    started_at: str
    completed_at: str = ""
    manufacturer: str = "Edmund Optics"
    source_directory: str = ""
    database_path: str = ""
    csv_path: str = ""
    accepted: int = 0
    incomplete: int = 0
    rejected: int = 0
    duplicates: int = 0
    files: list[dict[str, Any]] = field(default_factory=list)
    issues: list[dict[str, Any]] = field(default_factory=list)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def parse_float(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    match = NUMBER_PATTERN.search(clean_text(value).replace(",", ""))
    return float(match.group()) if match else None


def parse_wavelength_range(value: Any) -> tuple[float | None, float | None]:
    values = [float(item) for item in NUMBER_PATTERN.findall(clean_text(value))]
    if not values:
        return None, None
    if len(values) == 1:
        return values[0], values[0]
    return min(values[0], values[1]), max(values[0], values[1])


def split_materials(value: Any) -> list[str]:
    return [item.strip() for item in clean_text(value).split("/") if item.strip()]


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _surface(
    radius_mm: float | None,
    material_after: str,
    thickness_after_mm: float | None,
    clear_aperture_mm: float | None,
    coating: str,
) -> SurfaceSpec:
    return SurfaceSpec(
        radius_mm=radius_mm,
        material_after=material_after,
        thickness_after_mm=float(thickness_after_mm or 0.0),
        clear_aperture_mm=clear_aperture_mm,
        coating=coating,
        radius_locked=True,
        thickness_locked=True,
        material_locked=True,
        clear_aperture_locked=True,
    )


def parse_product(raw: dict[str, Any]) -> ParsedProduct:
    product_type = clean_text(raw.get("タイプ"))
    shape = SHAPE_MAP.get(product_type, "unknown")
    part_number = clean_text(raw.get("商品コード"))
    title = clean_text(raw.get("タイトル")) or part_number
    diameter = parse_float(raw.get("直径 (mm)"))
    clear_aperture = parse_float(raw.get("CA (mm)"))
    efl = parse_float(raw.get("EFL (mm)"))
    bfl = parse_float(raw.get("BFL (mm)"))
    coating = clean_text(raw.get("コーティング"))
    wavelength_min, wavelength_max = parse_wavelength_range(raw.get("波長範囲 (nm)"))
    reference_wavelength = parse_float(raw.get("焦点距離を規定している波長 (nm)"))
    materials = split_materials(raw.get("基板"))

    r1 = parse_float(raw.get("曲率半径 R1 (mm)"))
    r2 = parse_float(raw.get("曲率半径 R2 (mm)"))
    r3 = parse_float(raw.get("曲率半径 R3 (mm)"))
    symmetric_radius = parse_float(raw.get("曲率半径 R1=-R2 (mm)"))

    surfaces: list[SurfaceSpec] = []
    missing: list[str] = []

    if shape == "achromatic_doublet":
        ct1 = parse_float(raw.get("CT 1 (mm)"))
        ct2 = parse_float(raw.get("CT 2 (mm)"))
        if len(materials) != 2:
            missing.append("two verified glass names")
        if any(value is None for value in (r1, r2, r3, ct1, ct2)):
            missing.append("R1/R2/R3 and CT1/CT2")
        if len(materials) >= 2 and None not in (r1, r2, r3, ct1, ct2):
            surfaces = [
                _surface(r1, materials[0], ct1, clear_aperture, coating),
                _surface(r2, materials[1], ct2, clear_aperture, ""),
                _surface(r3, "air", 0.0, clear_aperture, coating),
            ]
    elif shape in {"plano_convex", "plano_concave", "double_convex", "double_concave"}:
        center_thickness = parse_float(raw.get("CT (mm)"))
        if len(materials) != 1:
            missing.append("one verified glass name")
        if center_thickness is None:
            missing.append("center thickness")

        if shape == "plano_convex":
            radius_first, radius_second = (abs(r1), None) if r1 is not None else (None, None)
            if r1 is None:
                missing.append("R1")
        elif shape == "plano_concave":
            radius_first, radius_second = (-abs(r1), None) if r1 is not None else (None, None)
            if r1 is None:
                missing.append("R1")
        elif shape == "double_convex":
            first = symmetric_radius if symmetric_radius is not None else r1
            second = symmetric_radius if symmetric_radius is not None else r2
            radius_first = abs(first) if first is not None else None
            radius_second = -abs(second) if second is not None else None
            if first is None or second is None:
                missing.append("R1/R2")
        else:
            first = symmetric_radius if symmetric_radius is not None else r1
            second = symmetric_radius if symmetric_radius is not None else r2
            radius_first = -abs(first) if first is not None else None
            radius_second = abs(second) if second is not None else None
            if first is None or second is None:
                missing.append("R1/R2")

        if len(materials) == 1 and center_thickness is not None and "R1" not in missing and "R1/R2" not in missing:
            surfaces = [
                _surface(radius_first, materials[0], center_thickness, clear_aperture, coating),
                _surface(radius_second, "air", 0.0, clear_aperture, coating),
            ]
    else:
        missing.append("supported product type")

    if not part_number:
        missing.append("part number")
    if diameter is None or diameter <= 0:
        missing.append("outer diameter")
    if clear_aperture is None or clear_aperture <= 0:
        missing.append("clear aperture")

    designable = not missing and bool(surfaces)
    return ParsedProduct(
        part_number=part_number,
        title=title,
        product_type=product_type,
        shape=shape,
        outer_diameter_mm=diameter,
        clear_aperture_mm=clear_aperture,
        effective_focal_length_mm=efl,
        back_focal_length_mm=bfl,
        coating=coating,
        wavelength_min_nm=wavelength_min,
        wavelength_max_nm=wavelength_max,
        reference_wavelength_nm=reference_wavelength,
        surfaces=surfaces,
        designable=designable,
        raw={key: clean_text(value) for key, value in raw.items()},
        missing_fields=sorted(set(missing)),
    )


def workbook_rows(path: Path) -> tuple[str, list[tuple[int, dict[str, Any]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(iterator)]
    rows: list[tuple[int, dict[str, Any]]] = []
    for source_row, values in enumerate(iterator, start=2):
        if not any(value is not None and clean_text(value) for value in values):
            continue
        rows.append((source_row, dict(zip(headers, values, strict=False))))
    workbook.close()
    return sheet.title, rows


def _write_csv(connection: sqlite3.Connection, path: Path) -> None:
    columns = [
        "manufacturer",
        "part_number",
        "title",
        "shape",
        "outer_diameter_mm",
        "clear_aperture_mm",
        "effective_focal_length_mm",
        "back_focal_length_mm",
        "coating",
        "wavelength_min_nm",
        "wavelength_max_nm",
        "reference_wavelength_nm",
        "designable",
        "r1_mm",
        "r2_mm",
        "r3_mm",
        "glass1",
        "glass2",
        "ct1_mm",
        "ct2_mm",
        "source_file",
        "source_sheet",
        "source_row",
        "source_sha256",
        "imported_at",
    ]
    rows = connection.execute(
        """
        SELECT p.*, m.name AS manufacturer, pr.source_file, pr.source_sheet,
               pr.source_row, pr.source_sha256, pr.retrieved_at,
               s0.radius_mm AS r1, s1.radius_mm AS r2, s2.radius_mm AS r3,
               s0.material_after AS glass1, s1.material_after AS medium2,
               s0.thickness_after_mm AS ct1, s1.thickness_after_mm AS ct2
        FROM products p
        JOIN manufacturers m ON m.id=p.manufacturer_id
        JOIN provenance pr ON pr.product_id=p.id
        LEFT JOIN surfaces s0 ON s0.product_id=p.id AND s0.surface_index=0
        LEFT JOIN surfaces s1 ON s1.product_id=p.id AND s1.surface_index=1
        LEFT JOIN surfaces s2 ON s2.product_id=p.id AND s2.surface_index=2
        ORDER BY m.name, p.part_number
        """
    ).fetchall()
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            has_third_surface = row["r3"] is not None
            writer.writerow(
                {
                    "manufacturer": row["manufacturer"],
                    "part_number": row["part_number"],
                    "title": row["title"],
                    "shape": row["shape"],
                    "outer_diameter_mm": row["outer_diameter_mm"],
                    "clear_aperture_mm": row["clear_aperture_mm"],
                    "effective_focal_length_mm": row["effective_focal_length_mm"],
                    "back_focal_length_mm": row["back_focal_length_mm"],
                    "coating": row["coating"],
                    "wavelength_min_nm": row["wavelength_min_nm"],
                    "wavelength_max_nm": row["wavelength_max_nm"],
                    "reference_wavelength_nm": row["reference_wavelength_nm"],
                    "designable": row["designable"],
                    "r1_mm": row["r1"],
                    "r2_mm": row["r2"],
                    "r3_mm": row["r3"],
                    "glass1": row["glass1"],
                    "glass2": row["medium2"] if has_third_surface else "",
                    "ct1_mm": row["ct1"],
                    "ct2_mm": row["ct2"] if has_third_surface else "",
                    "source_file": row["source_file"],
                    "source_sheet": row["source_sheet"],
                    "source_row": row["source_row"],
                    "source_sha256": row["source_sha256"],
                    "imported_at": row["retrieved_at"],
                }
            )


def import_edmund_catalog(
    source_directory: str | Path,
    database_path: str | Path,
    csv_path: str | Path,
    report_path: str | Path,
) -> ImportReport:
    source_directory = Path(source_directory)
    database_path = Path(database_path)
    csv_path = Path(csv_path)
    report_path = Path(report_path)
    started = datetime.now(UTC).isoformat()
    report = ImportReport(
        started_at=started,
        source_directory=str(source_directory.resolve()),
        database_path=str(database_path.resolve()),
        csv_path=str(csv_path.resolve()),
    )
    files = sorted(source_directory.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No .xlsx files found in {source_directory}")

    database_path.parent.mkdir(parents=True, exist_ok=True)
    if database_path.exists():
        database_path.unlink()

    connection = connect(database_path)
    try:
        initialize_schema(connection)
        manufacturer_id = connection.execute(
            "INSERT INTO manufacturers(name) VALUES (?) RETURNING id",
            (report.manufacturer,),
        ).fetchone()[0]
        run_id = connection.execute(
            "INSERT INTO import_runs(started_at, source_kind) VALUES (?, ?) RETURNING id",
            (started, "Edmund Optics Excel export"),
        ).fetchone()[0]

        for source_file in files:
            sheet_name, rows = workbook_rows(source_file)
            digest = file_sha256(source_file)
            file_summary = {
                "file": source_file.name,
                "sheet": sheet_name,
                "sha256": digest,
                "rows": len(rows),
                "accepted": 0,
                "incomplete": 0,
                "rejected": 0,
                "duplicates": 0,
            }
            for source_row, raw in rows:
                product = parse_product(raw)
                if not product.part_number:
                    report.rejected += 1
                    file_summary["rejected"] += 1
                    report.issues.append(
                        {"file": source_file.name, "row": source_row, "status": "rejected", "reason": "missing part number"}
                    )
                    continue
                existing = connection.execute(
                    "SELECT id FROM products WHERE manufacturer_id=? AND part_number=?",
                    (manufacturer_id, product.part_number),
                ).fetchone()
                if existing:
                    report.duplicates += 1
                    file_summary["duplicates"] += 1
                    report.issues.append(
                        {
                            "file": source_file.name,
                            "row": source_row,
                            "part_number": product.part_number,
                            "status": "duplicate",
                        }
                    )
                    continue

                cursor = connection.execute(
                    """
                    INSERT INTO products(
                        manufacturer_id, part_number, title, product_type, shape,
                        outer_diameter_mm, clear_aperture_mm, effective_focal_length_mm,
                        back_focal_length_mm, coating, wavelength_min_nm,
                        wavelength_max_nm, reference_wavelength_nm, designable, raw_json
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        manufacturer_id,
                        product.part_number,
                        product.title,
                        product.product_type,
                        product.shape,
                        product.outer_diameter_mm,
                        product.clear_aperture_mm,
                        product.effective_focal_length_mm,
                        product.back_focal_length_mm,
                        product.coating,
                        product.wavelength_min_nm,
                        product.wavelength_max_nm,
                        product.reference_wavelength_nm,
                        int(product.designable),
                        raw_json(product.raw),
                    ),
                )
                product_id = int(cursor.lastrowid)
                if product.surfaces:
                    insert_surfaces(connection, product_id, product.surfaces)
                connection.execute(
                    """
                    INSERT INTO provenance(
                        product_id, import_run_id, source_file, source_sheet,
                        source_row, source_sha256, retrieved_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (product_id, run_id, source_file.name, sheet_name, source_row, digest, started),
                )

                if product.designable:
                    report.accepted += 1
                    file_summary["accepted"] += 1
                else:
                    report.incomplete += 1
                    file_summary["incomplete"] += 1
                    report.issues.append(
                        {
                            "file": source_file.name,
                            "row": source_row,
                            "part_number": product.part_number,
                            "status": "incomplete",
                            "missing": product.missing_fields,
                        }
                    )
            report.files.append(file_summary)

        completed = datetime.now(UTC).isoformat()
        report.completed_at = completed
        connection.execute(
            """
            UPDATE import_runs
            SET completed_at=?, accepted_count=?, incomplete_count=?, rejected_count=?
            WHERE id=?
            """,
            (completed, report.accepted, report.incomplete, report.rejected, run_id),
        )
        connection.commit()
        _write_csv(connection, csv_path)
    finally:
        connection.close()

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(asdict(report), ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def default_paths() -> tuple[Path, Path, Path, Path]:
    repository = Path(__file__).resolve().parents[3]
    source = repository / "data" / "source" / "edmund"
    generated = repository / "data" / "generated"
    return (
        source,
        generated / "edmund_catalog.sqlite3",
        generated / "edmund_catalog.csv",
        generated / "edmund_import_report.json",
    )


def main(argv: Iterable[str] | None = None) -> int:
    source, database, csv_output, report = default_paths()
    parser = argparse.ArgumentParser(description="Import Edmund Optics Excel exports")
    parser.add_argument("--source", type=Path, default=source)
    parser.add_argument("--database", type=Path, default=database)
    parser.add_argument("--csv", type=Path, default=csv_output)
    parser.add_argument("--report", type=Path, default=report)
    args = parser.parse_args(list(argv) if argv is not None else None)
    result = import_edmund_catalog(args.source, args.database, args.csv, args.report)
    print(
        f"Imported {result.accepted} designable products; "
        f"{result.incomplete} incomplete, {result.rejected} rejected, {result.duplicates} duplicate"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
